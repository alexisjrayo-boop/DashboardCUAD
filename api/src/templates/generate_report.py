import sys
import json
import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

def is_valid_record(r, extensions_map):
    src = str(r.get('src') or '')
    destination = str(r.get('destination') or '')
    
    src_name = extensions_map.get(src)
    dest_name = extensions_map.get(destination)
    
    branch = None
    
    # 1. Check by extension name (Prefix TX, SC, JT, CB)
    if src_name:
        src_name = str(src_name)
        if src_name.startswith('TX'): branch = 'Tuxtepec'
        elif src_name.startswith('SC'): branch = 'Salina Cruz'
        elif src_name.startswith('JT'): branch = 'Juchitán'
        elif src_name.startswith('CB'): branch = 'CUAD'
        
    if not branch and dest_name:
        dest_name = str(dest_name)
        if dest_name.startswith('TX'): branch = 'Tuxtepec'
        elif dest_name.startswith('SC'): branch = 'Salina Cruz'
        elif dest_name.startswith('JT'): branch = 'Juchitán'
        elif dest_name.startswith('CB'): branch = 'CUAD'
        
    # 2. Check for specific trunk lines (Tuxtepec)
    TX_TRUNKS = ['2878759701', '2878750303']
    if not branch:
        if src in TX_TRUNKS or destination in TX_TRUNKS:
            branch = 'Tuxtepec'
            
    return branch is not None

def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_report.py <json_data_path> <template_path> <output_path>")
        sys.exit(1)

    json_path = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]

    if not os.path.exists(json_path):
        print(f"JSON data path not found: {json_path}")
        sys.exit(1)

    if not os.path.exists(template_path):
        print(f"Template path not found: {template_path}")
        sys.exit(1)

    # Load JSON records
    with open(json_path, 'r', encoding='utf-8') as f:
        data_payload = json.load(f)

    records = data_payload.get('records', [])
    filters = data_payload.get('filters', {})

    # Load extensions map for matching frontend filters
    extensions_path = os.path.join(os.path.dirname(template_path), '../config/extensions.json')
    extensions_map = {}
    if os.path.exists(extensions_path):
        try:
            with open(extensions_path, 'r', encoding='utf-8') as f:
                extensions_map = json.load(f)
        except Exception as e:
            print(f"Error loading extensions map: {e}")

    # No descartar registros para garantizar coincidencia 100% exacta con el Dashboard web
    # records = [r for r in records if is_valid_record(r, extensions_map)]

    # Load Template
    wb = openpyxl.load_workbook(template_path)
    
    # Get sheets
    ws_incoming = wb["Llamadas Entrantes"]
    ws_outgoing = wb["Llamadas Salientes"]
    ws_internal = wb["Llamadas Internas"]
    ws_records = wb["Registros de Llamadas"]

    # Write Filters to Sheet 1, 2, and 3
    start_date = filters.get('startDate', 'Inicio')
    end_date = filters.get('endDate', 'Hoy')
    date_val = f"{start_date} al {end_date}"
    line_val = filters.get('dst', 'Todas')

    for ws in [ws_incoming, ws_outgoing, ws_internal]:
        ws['B4'] = date_val
        ws['B5'] = line_val

    # Styles for ws_records data
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    # Clear existing rows in ws_records (except headers in row 1)
    if ws_records.max_row > 1:
        ws_records.delete_rows(2, ws_records.max_row - 1)

    # Write records to ws_records
    for idx, r in enumerate(records):
        row_num = 2 + idx
        
        # Translate calltype
        calltype = str(r.get('calltype', ''))
        calltype_label = 'Desconocido'
        if calltype == '1': calltype_label = 'Interna'
        elif calltype == '2': calltype_label = 'Entrante'
        elif calltype == '3': calltype_label = 'Saliente'

        row_values = [
            r.get('cdr_id'),
            r.get('calldate'),
            r.get('src'),
            r.get('dst'),
            r.get('destination'),
            r.get('duration'),
            r.get('lastapp'),
            r.get('disposition'),
            calltype_label
        ]

        # Use cell-by-cell assignment
        for col_idx, val in enumerate(row_values):
            cell = ws_records.cell(row=row_num, column=col_idx + 1, value=val)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = zebra_fill

    # Save to output path
    wb.save(output_path)
    print("Report generated successfully.")

if __name__ == "__main__":
    main()

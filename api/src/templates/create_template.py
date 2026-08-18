import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

def build_type_sheet(ws, title, calltype_name):
    ws.views.sheetView[0].showGridLines = True
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 5
    
    # Title formatting
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 40
    title_cell = ws['A1']
    title_cell.value = f"REPORTE DE {title.upper()}"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C3002F", end_color="C3002F", fill_type="solid") # Nissan Red
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Filters
    ws['A3'] = "Filtros Aplicados:"
    ws['A3'].font = Font(bold=True, size=11)
    
    ws['A4'] = "Rango de Fecha:"
    ws['A5'] = "Línea / Extensión:"
    ws['A6'] = "Total Registros (Tipo):"
    
    ws['B4'] = "" # To be filled by server
    ws['B5'] = "" # To be filled by server
    ws['B6'] = f"=COUNTIF('Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"
    
    for row in range(4, 7):
        ws[f'A{row}'].font = Font(name="Arial", size=10)
        ws[f'B{row}'].font = Font(name="Arial", size=10, bold=True)
    
    # Stats Table Headers
    ws['A8'] = "Métrica"
    ws['B8'] = "Valor"
    for col in ['A', 'B']:
        cell = ws[f'{col}8']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
        cell.border = thin_border
        
    # Formulas for Stats Table (pointing to Sheet 4 'Registros de Llamadas')
    stats_formulas = [
        ("Total Llamadas", f"=COUNTIF('Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"),
        ("Contestadas (ANSWERED)", f"=COUNTIFS('Registros de Llamadas'!H2:H10000, \"ANSWERED\", 'Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"),
        ("No Contestadas (NO ANSWER)", f"=COUNTIFS('Registros de Llamadas'!H2:H10000, \"NO ANSWER\", 'Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"),
        ("Ocupadas (BUSY)", f"=COUNTIFS('Registros de Llamadas'!H2:H10000, \"BUSY\", 'Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"),
        ("Fallidas (FAILED)", f"=COUNTIFS('Registros de Llamadas'!H2:H10000, \"FAILED\", 'Registros de Llamadas'!I2:I10000, \"{calltype_name}\")"),
        ("Porcentaje de Efectividad", "=IF(B9>0, B10/B9, 0)")
    ]
    
    for idx, (label, formula) in enumerate(stats_formulas):
        row_num = 9 + idx
        ws[f'A{row_num}'] = label
        ws[f'B{row_num}'] = formula
        
        ws[f'A{row_num}'].border = thin_border
        ws[f'B{row_num}'].border = thin_border
        
        # Zebra striping
        if idx % 2 == 1:
            bg_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            ws[f'A{row_num}'].fill = bg_fill
            ws[f'B{row_num}'].fill = bg_fill
            
    # Format B14 as percentage
    ws['B14'].number_format = '0.0%'

    # Add Pie Chart for Disposition
    pie = PieChart()
    pie.title = f"Distribución de Llamadas {title}"
    labels = Reference(ws, min_col=1, min_row=10, max_row=13)
    data = Reference(ws, min_col=2, min_row=9, max_row=13) # Include B9 'Valor' as header
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 15
    pie.height = 8.5
    ws.add_chart(pie, "D3")

def create_template():
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Llamadas Entrantes
    ws1 = wb.active
    ws1.title = "Llamadas Entrantes"
    build_type_sheet(ws1, "Llamadas Entrantes", "Entrante")
    
    # 2. Sheet 2: Llamadas Salientes
    ws2 = wb.create_sheet(title="Llamadas Salientes")
    build_type_sheet(ws2, "Llamadas Salientes", "Saliente")
    
    # 3. Sheet 3: Llamadas Internas
    ws3 = wb.create_sheet(title="Llamadas Internas")
    build_type_sheet(ws3, "Llamadas Internas", "Interna")
    
    # 4. Sheet 4: Registros de Llamadas (Todos los registros)
    ws4 = wb.create_sheet(title="Registros de Llamadas")
    ws4.views.sheetView[0].showGridLines = True
    
    headers = [
        "ID Llamada", "Fecha y Hora", "Origen (src)", "Destino (dst)", 
        "Extensión Destino", "Duración (seg)", "Acción", "Estado", "Tipo de Llamada"
    ]
    ws4.append(headers)
    
    # Style header row in Sheet 4
    ws4.row_dimensions[1].height = 25
    header_fill = PatternFill(start_color="C3002F", end_color="C3002F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")
        
    # Standard column widths for Sheet 4
    widths = [15, 22, 15, 15, 20, 15, 18, 15, 18]
    for idx, width in enumerate(widths):
        ws4.column_dimensions[get_column_letter(idx + 1)].width = width
        
    # Save template inside C:\Users\CB-AUXSISTEMAS\Desktop\DashboardCUAD\api\src\templates
    output_dir = r"C:\Users\CB-AUXSISTEMAS\Desktop\DashboardCUAD\api\src\templates"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    wb.save(os.path.join(output_dir, "report_template.xlsx"))
    print("4-Sheet Template created successfully at templates folder.")

if __name__ == "__main__":
    create_template()
